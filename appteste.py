import openpyxl

# Criar excel
book = openpyxl.Workbook()

# Nomear abas
book.create_sheet('Teste')
book.create_sheet('Teste2')
book.move_sheet('Sheet', 10000)

# Selecionar aba para preencher
Teste_page = book['Teste']

# Preencher linhas
Teste_page.append(['Nome', 'Sobrenome','Idade'])
Teste_page.append(['Bruno', 'Passareli','25'])
Teste_page.append(['Vitória', 'Dias','22'])
Teste_page.append(['Gustavo', 'Passareli','18'])

# Selecionar outra aba para preencher
Page_teste = book['Teste2']

# Preencher linhas
Page_teste.append(['Carro', 'Ano', 'Potência'])
Page_teste.append(['Polo TSI!', '2023', '1.0'])
Page_teste.append(['Civic LXL', '2011', '2.0'])
Page_teste.append(['Kwid ZEN', '2022', '1.0'])

# Salvar arquivo
book.save('Teste2.xlsx')




# --------------------------------------------------------

# Abrir o excel
book = openpyxl.load_workbook('Teste2.xlsx')

# Selecionar aba
Page_teste = book['Teste2']

# Exibir dados
for rows in Page_teste.iter_rows(min_row = 4):
    print(rows[0].value, rows[1].value, rows[2].value)
